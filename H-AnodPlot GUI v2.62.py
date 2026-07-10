# ==========================================
# H-AnodPlot GUI v2.62
# Anodização · Análise Eletroquímica
# Autor: Carlos Henrique Amaro da Silva
# ==========================================

import os
import numpy as np
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import tempfile
import re  # NOVO - para processar números

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook

# ─── RESOURCE PATH (PyInstaller) ───────────────────────────────────────────────
import sys

def resource_path(relative_path):
    """Resolve o caminho de recursos — funciona tanto em desenvolvimento
    quanto empacotado com PyInstaller (--onefile ou --onedir)."""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


# ─── PALETA ────────────────────────────────────────────────────────────────────
BG        = "#0f1117"
SURFACE   = "#1a1d27"
CARD      = "#22263a"
ACCENT    = "#4f8ef7"
ACCENT2   = "#7c3aed"
SUCCESS   = "#22c55e"
WARNING   = "#f59e0b"
ERROR     = "#ef4444"
TEXT      = "#e2e8f0"
TEXT_DIM  = "#64748b"
BORDER    = "#2d3148"

FONT_TITLE  = ("Segoe UI", 18, "bold")
FONT_HEAD   = ("Segoe UI", 11, "bold")
FONT_BODY   = ("Segoe UI", 10)
FONT_SMALL  = ("Segoe UI", 9)
FONT_MONO   = ("Consolas", 9)

ESTILOS_LINHA = {
    "Sólida":      "-",
    "Tracejada":   "--",
    "Pontilhada":  ":",
    "Traço-ponto": "-.",
}


# =============================================================================
# MODELO
# =============================================================================
CORES_TAB10 = [
    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
    "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf",
]

class Experimento:
    _contador = 0

    def __init__(self, nome, caminho):
        self.nome      = nome
        self.caminho   = caminho
        self.tempo     = []
        self.potencial = []
        self.densidade = []
        self.linestyle = "-"
        self.linewidth = 1.8
        # cor padrão sequencial do tab10
        idx = Experimento._contador % len(CORES_TAB10)
        self.color     = CORES_TAB10[idx]
        Experimento._contador += 1


class AnodPlot:
    def __init__(self):
        self.experimentos: list[Experimento] = []
        self.processado = False
        # rótulos editáveis — atualizados pelo gráfico interativo
        self.titulo_pot  = "Potencial vs Tempo"
        self.titulo_den  = "Densidade de Corrente vs Tempo"
        self.xlabel_pot  = "t (s)"
        self.ylabel_pot  = "V (V)"
        self.xlabel_den  = "t (s)"
        self.ylabel_den  = "i (mA/cm²)"
        self.mostrar_grid = True

    def carregar_arquivos(self, arquivos):
        self.experimentos = []
        Experimento._contador = 0
        for arq in arquivos:
            nome = os.path.basename(arq).replace(".dat", "")
            self.experimentos.append(Experimento(nome, arq))
        self.processado = False

    def aplicar_areas(self, areas):
        for exp, area in zip(self.experimentos, areas):
            tempo, potencial, densidade = [], [], []
            with open(exp.caminho, "r") as f:
                for linha in f:
                    linha = linha.split("*")[0].strip()
                    if not linha:
                        continue
                    partes = linha.split()
                    try:
                        t = float(partes[0])
                        V = float(partes[2].replace(",", "."))
                        I = float(partes[4].replace(",", "."))
                        j = (I * 1000) / area
                        tempo.append(t)
                        potencial.append(V)
                        densidade.append(j)
                    except Exception:
                        continue
            exp.tempo     = tempo
            exp.potencial = potencial
            exp.densidade = densidade
        self.processado = True

    def plotar(self):
        if not self.processado:
            return False

        # AUMENTA O TAMANHO DA FIGURA INTERATIVA
        fig, axes = plt.subplots(1, 2, figsize=(16, 6))
        fig.patch.set_facecolor("white")
        
        for ax in axes:
            ax.set_facecolor("white")
            ax.tick_params(colors="#111111", labelsize=11)
            ax.xaxis.label.set_color("#111111")
            ax.yaxis.label.set_color("#111111")
            ax.title.set_color("#111111")
            for sp in ax.spines.values():
                sp.set_edgecolor("#aaaaaa")
                sp.set_linewidth(1.0)

        nomes     = [e.nome for e in self.experimentos]
        line_refs = [[], []]

        for i, e in enumerate(self.experimentos):
            # LINHAS MAIS GROSSAS PARA VISUALIZAÇÃO
            lw = e.linewidth * 1.2
            l0, = axes[0].plot(e.tempo, e.potencial,
                               label=nomes[i], color=e.color,
                               linewidth=lw, linestyle=e.linestyle)
            l1, = axes[1].plot(e.tempo, e.densidade,
                               label=nomes[i], color=e.color,
                               linewidth=lw, linestyle=e.linestyle)
            line_refs[0].append(l0)
            line_refs[1].append(l1)

        # TAMANHOS DE FONTE MAIORES
        axes[0].set_xlabel(self.xlabel_pot, fontsize=12)
        axes[0].set_ylabel(self.ylabel_pot, fontsize=12)
        axes[0].set_title(self.titulo_pot, fontsize=13, fontweight="bold")

        axes[1].set_xlabel(self.xlabel_den, fontsize=12)
        axes[1].set_ylabel(self.ylabel_den, fontsize=12)
        axes[1].set_title(self.titulo_den, fontsize=13, fontweight="bold")

        for ax in axes:
            ax.legend(facecolor="white", edgecolor="#ccccca",
                      labelcolor="#111111", fontsize=10)
            if self.mostrar_grid:
                ax.grid(color="#dddddd", linestyle="--", linewidth=0.7)
            else:
                ax.grid(False)
            ax.title.set_backgroundcolor("none")
            ax.xaxis.label.set_backgroundcolor("none")
            ax.yaxis.label.set_backgroundcolor("none")

        fig.tight_layout(pad=2.5)
        fig.subplots_adjust(bottom=0.12)
        fig.text(
            0.5, 0.02,
            "Duplo clique em título, rótulo de eixo ou nome na legenda para editar",
            ha="center", fontsize=9, color="#64748b"
        )

        # ── popup Tkinter para edição ─────────────────────────────────────────
        def _popup_editar(titulo_janela, texto_atual, callback):
            popup = tk.Toplevel()
            popup.title(titulo_janela)
            popup.configure(bg="white")
            popup.resizable(False, False)
            popup.grab_set()
            popup.attributes("-topmost", True)
            popup.update_idletasks()
            pw, ph = 360, 110
            sw, sh = popup.winfo_screenwidth(), popup.winfo_screenheight()
            popup.geometry(f"{pw}x{ph}+{(sw-pw)//2}+{(sh-ph)//2}")

            tk.Label(popup, text=titulo_janela, font=("Segoe UI", 9, "bold"),
                     fg="#555555", bg="white").pack(anchor="w", padx=14, pady=(10, 2))

            var = tk.StringVar(value=texto_atual)
            entry = tk.Entry(popup, textvariable=var, font=("Segoe UI", 11),
                             bg="white", fg="#111111", insertbackground="#111111",
                             relief="flat", highlightthickness=1,
                             highlightcolor="#4f8ef7", highlightbackground="#cccccc")
            entry.pack(fill="x", padx=14, pady=4, ipady=5)
            entry.select_range(0, "end")
            entry.focus_set()

            def _confirmar(event=None):
                novo = var.get().strip()
                popup.destroy()
                if novo:
                    callback(novo)
                fig.canvas.draw_idle()

            def _cancelar(event=None):
                popup.destroy()

            bf = tk.Frame(popup, bg="white")
            bf.pack(anchor="e", padx=14, pady=(2, 10))
            tk.Button(bf, text="Cancelar", command=_cancelar,
                      bg="#f1f5f9", fg="#555555", font=("Segoe UI", 9),
                      relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left", padx=4)
            tk.Button(bf, text="  OK  ", command=_confirmar,
                      bg="#4f8ef7", fg="white", font=("Segoe UI", 9, "bold"),
                      relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left")

            entry.bind("<Return>", _confirmar)
            entry.bind("<Escape>", _cancelar)
            popup.wait_window()

        # ── detecção de duplo clique ──────────────────────────────────────────
        def _on_dblclick(event):
            if not event.dblclick:
                return

            renderer = fig.canvas.get_renderer()

            for ax_idx, ax in enumerate(axes):
                is_pot = (ax_idx == 0)

                t = ax.title
                if t.get_window_extent(renderer).contains(event.x, event.y):
                    def _set_titulo(novo, _t=t, _pot=is_pot):
                        _t.set_text(novo)
                        _t.set_backgroundcolor("none")
                        if _pot:
                            self.titulo_pot = novo
                        else:
                            self.titulo_den = novo
                        fig.canvas.draw_idle()
                    _popup_editar("Editar título", t.get_text(), _set_titulo)
                    return

                xl = ax.xaxis.label
                if xl.get_window_extent(renderer).contains(event.x, event.y):
                    def _set_xl(novo, _xl=xl, _pot=is_pot):
                        _xl.set_text(novo)
                        _xl.set_backgroundcolor("none")
                        if _pot:
                            self.xlabel_pot = novo
                        else:
                            self.xlabel_den = novo
                        fig.canvas.draw_idle()
                    _popup_editar("Editar rótulo — Eixo X", xl.get_text(), _set_xl)
                    return

                yl = ax.yaxis.label
                if yl.get_window_extent(renderer).contains(event.x, event.y):
                    def _set_yl(novo, _yl=yl, _pot=is_pot):
                        _yl.set_text(novo)
                        _yl.set_backgroundcolor("none")
                        if _pot:
                            self.ylabel_pot = novo
                        else:
                            self.ylabel_den = novo
                        fig.canvas.draw_idle()
                    _popup_editar("Editar rótulo — Eixo Y", yl.get_text(), _set_yl)
                    return

                leg = ax.get_legend()
                if leg:
                    for idx_t, txt in enumerate(leg.get_texts()):
                        if txt.get_window_extent(renderer).contains(event.x, event.y):
                            def _set_legenda(novo, _idx=idx_t):
                                nomes[_idx] = novo
                                self.experimentos[_idx].nome = novo
                                for ax_i, _ax in enumerate(axes):
                                    line_refs[ax_i][_idx].set_label(novo)
                                    old_leg = _ax.get_legend()
                                    if old_leg:
                                        old_leg.remove()
                                    _ax.legend(facecolor="white", edgecolor="#cccccc",
                                               labelcolor="#111111")
                                fig.canvas.draw_idle()
                            _popup_editar("Editar nome da amostra",
                                          txt.get_text(), _set_legenda)
                            return

        fig.canvas.mpl_connect("button_press_event", _on_dblclick)
        plt.close(fig)
        return True

    def gerar_imagens(self, pasta=None, temporario=False):
        """Gera imagens em alta resolução para exportação."""
        if temporario:
            pasta = tempfile.gettempdir()
        pot = os.path.join(pasta, "grafico_potencial.png")
        den = os.path.join(pasta, "grafico_densidade.png")

        for caminho, attr_y, ylabel_key, titulo_key, xlabel_key in [
            (pot, "potencial", "ylabel_pot", "titulo_pot", "xlabel_pot"),
            (den, "densidade", "ylabel_den", "titulo_den", "xlabel_den"),
        ]:
            # AUMENTA O TAMANHO DA FIGURA
            fig, ax = plt.subplots(figsize=(14, 8))
            fig.patch.set_facecolor("white")
            ax.set_facecolor("white")
            
            # AUMENTA O TAMANHO DA FONTE
            ax.tick_params(colors="#111111", labelsize=12)
            ax.xaxis.label.set_color("#111111")
            ax.yaxis.label.set_color("#111111")
            ax.title.set_color("#111111")
            
            # AUMENTA A ESPESSURA DAS LINHAS
            for sp in ax.spines.values():
                sp.set_edgecolor("#aaaaaa")
                sp.set_linewidth(1.2)
            
            # PLOT COM LINHAS MAIS GROSSAS PARA MELHOR RESOLUÇÃO
            for i, e in enumerate(self.experimentos):
                ax.plot(e.tempo, getattr(e, attr_y),
                        label=e.nome,
                        color=e.color,
                        linewidth=e.linewidth * 1.5,
                        linestyle=e.linestyle)
            
            # AUMENTA O TAMANHO DAS FONTES DOS RÓTULOS
            ax.set_xlabel(getattr(self, xlabel_key), fontsize=14)
            ax.set_ylabel(getattr(self, ylabel_key), fontsize=14)
            ax.set_title(getattr(self, titulo_key), fontsize=16, fontweight="bold")
            
            # AUMENTA O TAMANHO DA LEGENDA
            ax.legend(facecolor="white", edgecolor="#cccccc", labelcolor="#111111", 
                      fontsize=12)
            
            if self.mostrar_grid:
                ax.grid(color="#dddddd", linestyle="--", linewidth=0.8)
            
            fig.tight_layout()
            
            # SALVA COM DPI E ALTA QUALIDADE
            fig.savefig(caminho, dpi=300, facecolor="white", 
                       bbox_inches="tight", pad_inches=0.1)
            plt.close(fig)

        return pot, den

    def exportar_png(self, pasta):
        if not self.processado or not pasta:
            return False
        self.gerar_imagens(pasta=pasta)
        return True

    def exportar_excel(self, pasta):
        if not self.processado or not pasta:
            return False
        caminho_xlsx = os.path.join(pasta, "dados.xlsx")
        wb = Workbook()
        wb.remove(wb.active)

        for e in self.experimentos:
            ws = wb.create_sheet(title=e.nome[:31])
            ws.append(["Tempo (s)", "Potencial (V)", "Densidade de Corrente (mA/cm²)"])
            for t, v, j in zip(e.tempo, e.potencial, e.densidade):
                ws.append([t, v, j])

        ws_sum = wb.create_sheet(title="Resumo", index=0)
        ws_sum.append(["Experimento", "Pico J (mA/cm²)", "J Final (mA/cm²)",
                       "Pico V (V)", "V Final (V)"])
        for e in self.experimentos:
            j = np.array(e.densidade)
            v = np.array(e.potencial)
            ws_sum.append([e.nome, round(float(np.max(j)), 4), round(float(j[-1]), 4),
                           round(float(np.max(v)), 4), round(float(v[-1]), 4)])

        wb.save(caminho_xlsx)
        return caminho_xlsx

    def gerar_pdf(self, pasta):
        if not self.processado or not pasta:
            return False
        pot, den = self.gerar_imagens(temporario=True)
        caminho_pdf = os.path.join(pasta, "relatorio.pdf")
        doc = SimpleDocTemplate(caminho_pdf)
        styles = getSampleStyleSheet()
        elementos = []
        elementos.append(Paragraph("Relatório H-AnodPlot", styles["Title"]))
        elementos.append(Spacer(1, 12))
        for e in self.experimentos:
            j = np.array(e.densidade)
            v = np.array(e.potencial)
            texto = (
                f"<b>{e.nome}</b><br/>"
                f"Pico de densidade de corrente: {np.max(j):.3f} mA/cm²<br/>"
                f"Densidade de corrente final: {j[-1]:.3f} mA/cm²<br/>"
                f"Pico de potencial: {np.max(v):.3f} V<br/>"
                f"Potencial final: {v[-1]:.3f} V<br/><br/>"
            )
            elementos.append(Paragraph(texto, styles["Normal"]))
        # AUMENTA O TAMANHO DAS IMAGENS NO PDF
        elementos.append(Image(pot, width=500, height=300))
        elementos.append(Spacer(1, 12))
        elementos.append(Image(den, width=500, height=300))
        doc.build(elementos)
        return caminho_pdf


# =============================================================================
# WIDGETS AUXILIARES
# =============================================================================
def btn(parent, text, command, color=ACCENT, width=None):
    kw = dict(
        text=text, command=command,
        bg=color, fg="white", activebackground=color,
        activeforeground="white",
        font=FONT_HEAD, relief="flat", cursor="hand2",
        padx=14, pady=8,
    )
    if width:
        kw["width"] = width
    b = tk.Button(parent, **kw)

    def on_enter(e):
        b.config(bg=_lighten(color))

    def on_leave(e):
        b.config(bg=color)

    b.bind("<Enter>", on_enter)
    b.bind("<Leave>", on_leave)
    return b


def _lighten(hex_color, amount=30):
    hex_color = hex_color.lstrip("#")
    r, g, b_ = (int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    return f"#{min(255,r+amount):02x}{min(255,g+amount):02x}{min(255,b_+amount):02x}"


def label(parent, text, font=FONT_BODY, fg=TEXT, **kw):
    return tk.Label(parent, text=text, font=font, fg=fg, bg=parent["bg"], **kw)


def card(parent, **kw):
    return tk.Frame(parent, bg=CARD, bd=0, highlightthickness=1,
                    highlightbackground=BORDER, **kw)


def separator(parent):
    return tk.Frame(parent, bg=BORDER, height=1)


# =============================================================================
# CANVAS EMBUTIDO - grafico dentro da janela principal
# =============================================================================
class GraficoEmbutidoAnod:
    def __init__(self, parent_frame, model: AnodPlot):
        self.model = model
        self.parent = parent_frame
        self.axes = []
        self.line_refs = [[], []]
        self._cid_dblclick = None

        self.fig = Figure(figsize=(9, 4.8), dpi=96)
        self.fig.patch.set_facecolor("white")

        self.canvas = FigureCanvasTkAgg(self.fig, master=parent_frame)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)

        toolbar_frame = tk.Frame(parent_frame, bg="white")
        toolbar_frame.pack(fill="x")
        self.toolbar = NavigationToolbar2Tk(self.canvas, toolbar_frame)
        self.toolbar.config(bg="white")
        self.toolbar.update()

        self._mostrar_placeholder()

    def _estilizar(self, ax):
        ax.set_facecolor("white")
        ax.tick_params(colors="#111111", labelsize=9)
        ax.xaxis.label.set_color("#111111")
        ax.yaxis.label.set_color("#111111")
        ax.title.set_color("#111111")
        for sp in ax.spines.values():
            sp.set_visible(True)
            sp.set_edgecolor("#aaaaaa")
            sp.set_linewidth(0.8)

    def _mostrar_placeholder(self):
        self.fig.clear()
        ax = self.fig.add_subplot(111)
        self._estilizar(ax)
        ax.text(0.5, 0.5,
                "Selecione arquivos .dat,\ninforme as areas e processe os dados",
                ha="center", va="center",
                transform=ax.transAxes,
                fontsize=12, color="#aaaaaa",
                fontfamily="Segoe UI")
        ax.set_xticks([])
        ax.set_yticks([])
        for sp in ax.spines.values():
            sp.set_visible(False)
        self.axes = [ax]
        self.canvas.draw_idle()

    def redesenhar(self):
        if not self.model.processado:
            self._mostrar_placeholder()
            return False

        self.fig.clear()
        axes = self.fig.subplots(1, 2)
        self.axes = list(axes)
        self.line_refs = [[], []]
        nomes = [e.nome for e in self.model.experimentos]

        for ax in self.axes:
            self._estilizar(ax)

        for i, e in enumerate(self.model.experimentos):
            l0, = self.axes[0].plot(e.tempo, e.potencial,
                                    label=nomes[i], color=e.color,
                                    linewidth=e.linewidth, linestyle=e.linestyle)
            l1, = self.axes[1].plot(e.tempo, e.densidade,
                                    label=nomes[i], color=e.color,
                                    linewidth=e.linewidth, linestyle=e.linestyle)
            self.line_refs[0].append(l0)
            self.line_refs[1].append(l1)

        self.axes[0].set_xlabel(self.model.xlabel_pot, fontsize=10)
        self.axes[0].set_ylabel(self.model.ylabel_pot, fontsize=10)
        self.axes[0].set_title(self.model.titulo_pot, fontsize=11, fontweight="bold")

        self.axes[1].set_xlabel(self.model.xlabel_den, fontsize=10)
        self.axes[1].set_ylabel(self.model.ylabel_den, fontsize=10)
        self.axes[1].set_title(self.model.titulo_den, fontsize=11, fontweight="bold")

        for ax in self.axes:
            ax.legend(facecolor="white", edgecolor="#cccccc",
                      labelcolor="#111111", fontsize=9, framealpha=0.9)
            if self.model.mostrar_grid:
                ax.grid(True, color="#dddddd", linestyle="--", linewidth=0.6)
            else:
                ax.grid(False)
            ax.title.set_backgroundcolor("none")
            ax.xaxis.label.set_backgroundcolor("none")
            ax.yaxis.label.set_backgroundcolor("none")

        self.fig.tight_layout(pad=2)
        self.fig.subplots_adjust(bottom=0.15)
        self.fig.text(
            0.5, 0.03,
            "Duplo clique em titulo, rotulo de eixo ou nome na legenda para editar",
            ha="center", fontsize=8, color="#64748b"
        )

        if self._cid_dblclick is not None:
            self.canvas.mpl_disconnect(self._cid_dblclick)
        self._cid_dblclick = self.canvas.mpl_connect(
            "button_press_event", self._on_dblclick
        )
        self.canvas.draw_idle()
        return True

    def _popup(self, root, titulo_janela, texto_atual, callback):
        popup = tk.Toplevel(root)
        popup.title(titulo_janela)
        popup.configure(bg="white")
        popup.resizable(False, False)
        popup.grab_set()
        popup.attributes("-topmost", True)
        popup.update_idletasks()
        pw, ph = 380, 115
        sw, sh = popup.winfo_screenwidth(), popup.winfo_screenheight()
        popup.geometry(f"{pw}x{ph}+{(sw-pw)//2}+{(sh-ph)//2}")

        tk.Label(popup, text=titulo_janela, font=("Segoe UI", 9, "bold"),
                 fg="#555555", bg="white").pack(anchor="w", padx=14, pady=(10, 2))

        var = tk.StringVar(value=texto_atual)
        entry = tk.Entry(popup, textvariable=var, font=("Segoe UI", 11),
                         bg="white", fg="#111111", insertbackground="#111111",
                         relief="flat", highlightthickness=1,
                         highlightcolor="#4f8ef7", highlightbackground="#cccccc")
        entry.pack(fill="x", padx=14, pady=4, ipady=5)
        entry.select_range(0, "end")
        entry.focus_set()

        def _confirmar(event=None):
            novo = var.get().strip()
            popup.destroy()
            if novo:
                callback(novo)
            self.canvas.draw_idle()

        def _cancelar(event=None):
            popup.destroy()

        bf = tk.Frame(popup, bg="white")
        bf.pack(anchor="e", padx=14, pady=(2, 10))
        tk.Button(bf, text="Cancelar", command=_cancelar,
                  bg="#f1f5f9", fg="#555555", font=("Segoe UI", 9),
                  relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left", padx=4)
        tk.Button(bf, text="  OK  ", command=_confirmar,
                  bg="#4f8ef7", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10, pady=4, cursor="hand2").pack(side="left")

        entry.bind("<Return>", _confirmar)
        entry.bind("<Escape>", _cancelar)
        popup.wait_window()

    def _on_dblclick(self, event):
        if not event.dblclick or not self.model.processado:
            return

        renderer = self.canvas.get_renderer()
        root = self.parent.winfo_toplevel()

        for ax_idx, ax in enumerate(self.axes[:2]):
            is_pot = (ax_idx == 0)

            title = ax.title
            if title.get_window_extent(renderer).contains(event.x, event.y):
                def _set_titulo(novo, _title=title, _pot=is_pot):
                    _title.set_text(novo)
                    if _pot:
                        self.model.titulo_pot = novo
                    else:
                        self.model.titulo_den = novo
                self._popup(root, "Editar titulo", title.get_text(), _set_titulo)
                return

            xlabel = ax.xaxis.label
            if xlabel.get_window_extent(renderer).contains(event.x, event.y):
                def _set_xlabel(novo, _xlabel=xlabel, _pot=is_pot):
                    _xlabel.set_text(novo)
                    if _pot:
                        self.model.xlabel_pot = novo
                    else:
                        self.model.xlabel_den = novo
                self._popup(root, "Editar rotulo - Eixo X", xlabel.get_text(), _set_xlabel)
                return

            ylabel = ax.yaxis.label
            if ylabel.get_window_extent(renderer).contains(event.x, event.y):
                def _set_ylabel(novo, _ylabel=ylabel, _pot=is_pot):
                    _ylabel.set_text(novo)
                    if _pot:
                        self.model.ylabel_pot = novo
                    else:
                        self.model.ylabel_den = novo
                self._popup(root, "Editar rotulo - Eixo Y", ylabel.get_text(), _set_ylabel)
                return

            legend = ax.get_legend()
            if legend:
                for idx_t, txt in enumerate(legend.get_texts()):
                    if txt.get_window_extent(renderer).contains(event.x, event.y):
                        def _set_legenda(novo, _idx=idx_t):
                            self.model.experimentos[_idx].nome = novo
                            for ax_i, axis in enumerate(self.axes[:2]):
                                self.line_refs[ax_i][_idx].set_label(novo)
                                old_leg = axis.get_legend()
                                if old_leg:
                                    old_leg.remove()
                                axis.legend(facecolor="white", edgecolor="#cccccc",
                                            labelcolor="#111111", fontsize=9,
                                            framealpha=0.9)
                        self._popup(root, "Editar nome da amostra",
                                    txt.get_text(), _set_legenda)
                        return

    def editar_titulo_pot(self, root):
        def _set(novo):
            self.model.titulo_pot = novo
            self.redesenhar()
        self._popup(root, "Editar titulo - Potencial", self.model.titulo_pot, _set)

    def editar_titulo_den(self, root):
        def _set(novo):
            self.model.titulo_den = novo
            self.redesenhar()
        self._popup(root, "Editar titulo - Densidade", self.model.titulo_den, _set)

    def editar_xlabel_pot(self, root):
        def _set(novo):
            self.model.xlabel_pot = novo
            self.redesenhar()
        self._popup(root, "Editar eixo X - Potencial", self.model.xlabel_pot, _set)

    def editar_ylabel_pot(self, root):
        def _set(novo):
            self.model.ylabel_pot = novo
            self.redesenhar()
        self._popup(root, "Editar eixo Y - Potencial", self.model.ylabel_pot, _set)

    def editar_xlabel_den(self, root):
        def _set(novo):
            self.model.xlabel_den = novo
            self.redesenhar()
        self._popup(root, "Editar eixo X - Densidade", self.model.xlabel_den, _set)

    def editar_ylabel_den(self, root):
        def _set(novo):
            self.model.ylabel_den = novo
            self.redesenhar()
        self._popup(root, "Editar eixo Y - Densidade", self.model.ylabel_den, _set)

    def editar_nome_amostra(self, root, idx):
        if idx is None or idx >= len(self.model.experimentos):
            return
        exp = self.model.experimentos[idx]
        def _set(novo):
            exp.nome = novo
            self.redesenhar()
        self._popup(root, "Editar nome da amostra", exp.nome, _set)


# =============================================================================
# TELA PRINCIPAL
# =============================================================================
class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("H-AnodPlot v2.62")
        try:
            self.root.iconbitmap(resource_path("AnodPlot.ico"))
        except Exception:
            pass  # ícone não encontrado — continua sem ele
        self.root.configure(bg=BG)
        self.root.resizable(True, True)
        self.root.minsize(1100, 700)

        self.app            = AnodPlot()
        self.arquivos_sel   : tuple       = ()
        self.pasta_saida    : str         = ""
        self.entradas_area  : list[tk.Entry] = []
        self.amostra_idx_sel: int | None  = None
        self.lista_rows     : list[tk.Frame] = []

        self._build()
        self.root.mainloop()

    # ── layout ────────────────────────────────────────────────────────────────
    def _build(self):
        header = tk.Frame(self.root, bg=SURFACE, pady=16)
        header.pack(fill="x")

        tk.Label(header, text="H-AnodPlot",
                 font=FONT_TITLE, fg=ACCENT, bg=SURFACE).pack(side="left", padx=24)
        tk.Label(header, text="Anodização  ·  Análise Eletroquímica",
                 font=FONT_SMALL, fg=TEXT_DIM, bg=SURFACE).pack(side="left", padx=4)
        tk.Label(header, text="Autor: Carlos Henrique Amaro da Silva",
                 font=FONT_SMALL, fg=TEXT_DIM, bg=SURFACE).pack(side="right", padx=24)

        body = tk.Frame(self.root, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)
        body.columnconfigure(0, weight=0)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        self._build_pipeline(body)
        self._build_workspace(body)

    def _build_pipeline(self, parent):
        # coluna esquerda com scroll para caber todos os elementos
        outer = tk.Frame(parent, bg=BG, width=330)
        outer.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        outer.grid_propagate(False)
        outer.rowconfigure(0, weight=1)
        outer.columnconfigure(0, weight=1)

        canvas_left = tk.Canvas(outer, bg=BG, bd=0, highlightthickness=0)
        canvas_left.grid(row=0, column=0, sticky="nsew")

        sb_left = tk.Scrollbar(outer, orient="vertical", command=canvas_left.yview)
        sb_left.grid(row=0, column=1, sticky="ns")
        canvas_left.configure(yscrollcommand=sb_left.set)

        col = tk.Frame(canvas_left, bg=BG)
        win_id = canvas_left.create_window((0, 0), window=col, anchor="nw")

        def _on_col_configure(e):
            canvas_left.configure(scrollregion=canvas_left.bbox("all"))
        def _on_canvas_left_configure(e):
            canvas_left.itemconfig(win_id, width=e.width)

        col.bind("<Configure>", _on_col_configure)
        canvas_left.bind("<Configure>", _on_canvas_left_configure)

        # scroll com roda do mouse apenas quando o cursor esta no painel esquerdo
        def _on_mousewheel(e):
            canvas_left.yview_scroll(int(-1*(e.delta/120)), "units")
        canvas_left.bind("<Enter>", lambda e: canvas_left.bind_all("<MouseWheel>", _on_mousewheel))
        canvas_left.bind("<Leave>", lambda e: canvas_left.unbind_all("<MouseWheel>"))
        col.bind("<Enter>", lambda e: canvas_left.bind_all("<MouseWheel>", _on_mousewheel))
        col.bind("<Leave>", lambda e: canvas_left.unbind_all("<MouseWheel>"))

        label(col, "FLUXO DE TRABALHO", font=("Segoe UI", 8, "bold"),
              fg=TEXT_DIM).pack(anchor="w", pady=(0, 10))

        steps = [
            ("1", "Selecionar arquivos",  "Escolha os arquivos .dat",        self._step1),
            ("2", "Definir pasta de saída","Onde salvar os resultados",       self._step2),
            ("3", "Informar Áreas",        "Área de cada eletrodo (cm²)",     None),
            ("4", "Processar dados",       "Aplicar normalização por Área",   self._step4),
        ]

        self.step_labels: dict[str, tk.Label] = {}

        for num, titulo, descricao, cmd in steps:
            c = card(col)
            c.pack(fill="x", pady=4, ipady=6)

            header_row = tk.Frame(c, bg=CARD)
            header_row.pack(fill="x", padx=12, pady=(8, 2))

            tk.Label(header_row, text=num, font=("Segoe UI", 9, "bold"),
                     fg="white", bg=ACCENT2, width=2,
                     relief="flat", padx=4, pady=2).pack(side="left")
            tk.Label(header_row, text=f"  {titulo}", font=FONT_HEAD,
                     fg=TEXT, bg=CARD).pack(side="left")
            tk.Label(c, text=descricao, font=FONT_SMALL, fg=TEXT_DIM,
                     bg=CARD).pack(anchor="w", padx=12, pady=(0, 6))

            status_lbl = tk.Label(c, text="Pendente", font=FONT_SMALL,
                                  fg=TEXT_DIM, bg=CARD)
            status_lbl.pack(anchor="w", padx=12, pady=(0, 4))
            self.step_labels[num] = status_lbl

            if cmd:
                btn(c, f"Executar passo {num}", cmd, color=ACCENT2).pack(
                    anchor="w", padx=12, pady=(0, 8))

            if num == "3":
                label(col, "ÁREAS DOS ELETRODOS (cm²)", font=("Segoe UI", 8, "bold"),
                      fg=TEXT_DIM).pack(anchor="w", pady=(6, 6))

                areas_card = card(col)
                areas_card.pack(fill="x", pady=(0, 8))

                # NOVO: Botão para colar áreas do Excel
                btn_areas_frame = tk.Frame(areas_card, bg=CARD)
                btn_areas_frame.pack(fill="x", padx=8, pady=(8, 4))
                
                btn(btn_areas_frame, "📋 Colar áreas do Excel (Ctrl+V)", 
                    self._colar_areas_excel, color="#0e7490").pack(side="left", padx=0)
                
                tk.Label(btn_areas_frame, 
                         text="Cole valores separados por TAB, espaço ou vírgula",
                         font=("Segoe UI", 7), fg=TEXT_DIM, bg=CARD).pack(side="left", padx=10)

                areas_outer = tk.Frame(areas_card, bg=CARD)
                areas_outer.pack(fill="x", padx=8, pady=(0, 8))

                self.scroll_canvas = tk.Canvas(areas_outer, bg=SURFACE,
                                               bd=0, highlightthickness=0, height=220)
                self.scroll_canvas.pack(side="left", fill="both", expand=True)

                sb = tk.Scrollbar(areas_outer, orient="vertical",
                                  command=self.scroll_canvas.yview)
                sb.pack(side="right", fill="y")
                self.scroll_canvas.configure(yscrollcommand=sb.set)

                self.areas_frame = tk.Frame(self.scroll_canvas, bg=SURFACE)
                self._canvas_window = self.scroll_canvas.create_window(
                    (0, 0), window=self.areas_frame, anchor="nw"
                )
                self.areas_frame.bind("<Configure>", self._on_frame_configure)
                self.scroll_canvas.bind("<Configure>", self._on_canvas_configure)
                self._ativar_scroll_areas()

        separator(col).pack(fill="x", pady=10)

        label(col, "VISUALIZACAO", font=("Segoe UI", 8, "bold"),
              fg=TEXT_DIM).pack(anchor="w", pady=(0, 8))

        btn(col, "Visualizar gráficos",      self._plotar,         color=ACCENT    ).pack(fill="x", pady=3)

        separator(col).pack(fill="x", pady=8)

        label(col, "OPÇÕES DO GRÁFICO", font=("Segoe UI", 8, "bold"),
              fg=TEXT_DIM).pack(anchor="w", pady=(0, 6))

        self.btn_grid = btn(col, "Grid  ON", self._toggle_grid, color="#0e7490")
        self.btn_grid.pack(fill="x", pady=3)

        separator(col).pack(fill="x", pady=8)

        label(col, "ESTILO DA AMOSTRA", font=("Segoe UI", 8, "bold"),
              fg=TEXT_DIM).pack(anchor="w", pady=(0, 6))

        estilo_card = card(col)
        estilo_card.pack(fill="x", pady=(0, 4))

        sel_row = tk.Frame(estilo_card, bg=CARD)
        sel_row.pack(fill="x", padx=10, pady=(8, 3))

        tk.Label(sel_row, text="Amostra:", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=CARD, width=9, anchor="w").pack(side="left")

        self.amostra_var = tk.StringVar(value="—")
        self.amostra_menu = tk.OptionMenu(sel_row, self.amostra_var, "—",
                                          command=self._on_amostra_select)
        self.amostra_menu.config(bg=BG, fg=TEXT, activebackground=ACCENT2,
                                 activeforeground="white", font=FONT_SMALL,
                                 relief="flat", highlightthickness=0, bd=0, width=17)
        self.amostra_menu["menu"].config(bg=BG, fg=TEXT, activebackground=ACCENT2,
                                         activeforeground="white", font=FONT_SMALL)
        self.amostra_menu.pack(side="left", padx=4)

        tk.Label(estilo_card, text="↑ escolha a amostra: cor, estilo e espessura abaixo valem só para ela",
                 font=("Segoe UI", 7), fg=TEXT_DIM, bg=CARD,
                 wraplength=290, justify="left").pack(anchor="w", padx=10, pady=(0, 2))

        tipo_row = tk.Frame(estilo_card, bg=CARD)
        tipo_row.pack(fill="x", padx=10, pady=3)

        tk.Label(tipo_row, text="Linha:", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=CARD, width=9, anchor="w").pack(side="left")

        self.estilo_var = tk.StringVar(value="Sólida")
        estilo_menu = tk.OptionMenu(tipo_row, self.estilo_var,
                                    *ESTILOS_LINHA.keys(),
                                    command=self._on_estilo_change)
        estilo_menu.config(bg=BG, fg=TEXT, activebackground=ACCENT2,
                           activeforeground="white", font=FONT_SMALL,
                           relief="flat", highlightthickness=0, bd=0, width=12)
        estilo_menu["menu"].config(bg=BG, fg=TEXT, activebackground=ACCENT2,
                                   activeforeground="white", font=FONT_SMALL)
        estilo_menu.pack(side="left", padx=4)

        cor_row = tk.Frame(estilo_card, bg=CARD)
        cor_row.pack(fill="x", padx=10, pady=3)

        tk.Label(cor_row, text="Cor:", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=CARD, width=9, anchor="w").pack(side="left")

        self.cor_preview = tk.Label(cor_row, bg=CORES_TAB10[0], width=4,
                                    relief="flat", cursor="hand2")
        self.cor_preview.pack(side="left", padx=4, ipady=7)
        self.cor_preview.bind("<Button-1>", lambda e: self._escolher_cor())

        tk.Label(cor_row, text="clicar (só esta amostra)", font=("Segoe UI", 8),
                 fg=TEXT_DIM, bg=CARD).pack(side="left", padx=2)

        lw_row = tk.Frame(estilo_card, bg=CARD)
        lw_row.pack(fill="x", padx=10, pady=(3, 8))

        tk.Label(lw_row, text="Espessura:", font=FONT_SMALL,
                 fg=TEXT_DIM, bg=CARD, width=9, anchor="w").pack(side="left")

        self.lw_var = tk.DoubleVar(value=1.8)
        tk.Scale(lw_row, from_=0.5, to=5.0, resolution=0.5,
                 orient="horizontal", variable=self.lw_var,
                 bg=CARD, fg=TEXT, troughcolor=BG,
                 highlightthickness=0, activebackground=ACCENT,
                 length=125, command=self._on_lw_change).pack(side="left")

        self.lw_label = tk.Label(lw_row, text="1.8 pt", font=FONT_SMALL,
                                  fg=TEXT_DIM, bg=CARD, width=5)
        self.lw_label.pack(side="left", padx=2)

        separator(col).pack(fill="x", pady=8)

        label(col, "EDITAR TEXTOS", font=("Segoe UI", 8, "bold"),
              fg=TEXT_DIM).pack(anchor="w", pady=(0, 6))

        btn(col, "Titulo - Potencial", self._editar_titulo_pot, color="#374151").pack(fill="x", pady=2)
        btn(col, "Titulo - Densidade", self._editar_titulo_den, color="#374151").pack(fill="x", pady=2)
        btn(col, "Eixo X - Potencial", self._editar_xlabel_pot, color="#374151").pack(fill="x", pady=2)
        btn(col, "Eixo Y - Potencial", self._editar_ylabel_pot, color="#374151").pack(fill="x", pady=2)
        btn(col, "Eixo X - Densidade", self._editar_xlabel_den, color="#374151").pack(fill="x", pady=2)
        btn(col, "Eixo Y - Densidade", self._editar_ylabel_den, color="#374151").pack(fill="x", pady=2)
        btn(col, "Nome da amostra", self._editar_nome_amostra, color="#374151").pack(fill="x", pady=2)

        separator(col).pack(fill="x", pady=10)

        label(col, "EXPORTAR", font=("Segoe UI", 8, "bold"),
              fg=TEXT_DIM).pack(anchor="w", pady=(0, 8))

        btn(col, "Exportar PNG",           self._exportar_png,   color="#0e7490" ).pack(fill="x", pady=3)
        btn(col, "Gerar relatorio PDF",    self._gerar_pdf,      color="#0e7490" ).pack(fill="x", pady=3)
        btn(col, "Exportar Excel (.xlsx)", self._exportar_excel, color="#166534" ).pack(fill="x", pady=3)

    def _build_workspace(self, parent):
        col = tk.Frame(parent, bg=BG)
        col.grid(row=0, column=1, sticky="nsew")

        # variáveis mantidas para compatibilidade com _sincronizar_painel e _atualizar_lista
        self.lista_frame   = tk.Frame(col, bg=BG)   # frame oculto — não empacotado
        self.info_var      = tk.StringVar(value="")  # variável sem widget visível

        grafico_card = card(col)
        grafico_card.pack(fill="both", expand=True)

        self.grafico = GraficoEmbutidoAnod(grafico_card, self.app)

        self.status_var = tk.StringVar(value="Pronto. Selecione arquivos .dat para comecar.")
        status_bar = tk.Frame(self.root, bg=SURFACE, pady=4)
        status_bar.pack(fill="x", side="bottom")
        tk.Label(status_bar, textvariable=self.status_var,
                 font=FONT_SMALL, fg=TEXT_DIM, bg=SURFACE).pack(side="left", padx=16)

    def _on_frame_configure(self, event):
        self.scroll_canvas.configure(scrollregion=self.scroll_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.scroll_canvas.itemconfig(self._canvas_window, width=event.width)

    def _ativar_scroll_areas(self):
        if not hasattr(self, "scroll_canvas"):
            return

        def _on_mousewheel(e):
            self.scroll_canvas.yview_scroll(int(-1*(e.delta/120)), "units")

        def _bind(_event=None):
            self.scroll_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind(_event=None):
            self.scroll_canvas.unbind_all("<MouseWheel>")

        for widget in (self.scroll_canvas, self.areas_frame):
            widget.bind("<Enter>", _bind)
            widget.bind("<Leave>", _unbind)

    def _vincular_scroll_areas_widget(self, widget):
        if not hasattr(self, "scroll_canvas"):
            return

        def _on_mousewheel(e):
            self.scroll_canvas.yview_scroll(int(-1*(e.delta/120)), "units")

        widget.bind("<Enter>", lambda e: self.scroll_canvas.bind_all("<MouseWheel>", _on_mousewheel))
        widget.bind("<Leave>", lambda e: self.scroll_canvas.unbind_all("<MouseWheel>"))

    # ── NOVO: Colar áreas do Excel ────────────────────────────────────────────
    def _colar_areas_excel(self):
        """Abre uma caixa de diálogo para colar múltiplas áreas de uma vez."""
        if not self.arquivos_sel:
            self._erro("Carregue os arquivos .dat primeiro (Passo 1).")
            return
        
        # Cria popup para colar os dados
        popup = tk.Toplevel(self.root)
        popup.title("Colar Áreas do Excel")
        popup.configure(bg="white")
        popup.resizable(True, True)
        popup.grab_set()
        popup.attributes("-topmost", True)
        
        pw, ph = 500, 350
        sw, sh = popup.winfo_screenwidth(), popup.winfo_screenheight()
        popup.geometry(f"{pw}x{ph}+{(sw-pw)//2}+{(sh-ph)//2}")
        
        tk.Label(popup, text="Cole os valores das áreas (uma por linha ou separados por TAB/espaço/vírgula)",
                 font=("Segoe UI", 10, "bold"), fg="#1e293b", bg="white").pack(pady=(12, 4))
        
        tk.Label(popup, 
                 text=f"Espera {len(self.arquivos_sel)} valores (um para cada arquivo carregado).\n"
                      "Exemplo: 2.5  3.0  1.8  2.2\n"
                      "Ou: 2.5, 3.0, 1.8, 2.2\n"
                      "Ou coluna do Excel: 2.5 ↵ 3.0 ↵ 1.8 ↵ 2.2",
                 font=("Segoe UI", 8), fg="#64748b", bg="white", justify="left").pack(pady=(0, 8))
        
        # Área de texto para colar
        text_area = tk.Text(popup, height=8, width=50, font=("Consolas", 11),
                            bg="#f8fafc", fg="#1e293b", relief="solid", 
                            highlightthickness=1, highlightcolor="#4f8ef7")
        text_area.pack(padx=14, pady=4, fill="both", expand=True)
        text_area.focus_set()
        
        # Frame dos botões
        btn_frame = tk.Frame(popup, bg="white")
        btn_frame.pack(pady=10)
        
        def _processar_areas():
            texto = text_area.get("1.0", "end").strip()
            if not texto:
                messagebox.showwarning("H-AnodPlot", "Nenhum dado foi colado.")
                return
            
            # Extrai todos os números do texto
            numeros = re.findall(r"[-+]?\d*\.?\d+", texto.replace(",", "."))
            
            if not numeros:
                messagebox.showerror("H-AnodPlot", 
                    "Nenhum número válido encontrado.\n"
                    "Certifique-se de colar valores numéricos.")
                return
            
            # Converte para float
            valores = [float(n) for n in numeros]
            
            # Verifica se a quantidade de valores bate com o número de arquivos
            n_arquivos = len(self.arquivos_sel)
            if len(valores) != n_arquivos:
                resposta = messagebox.askyesno(
                    "H-AnodPlot",
                    f"Foram encontrados {len(valores)} valores, mas há {n_arquivos} arquivos.\n"
                    f"Os valores excedentes serão ignorados.\n\n"
                    f"Deseja continuar?"
                )
                if not resposta:
                    return
                # Trunca para o número de arquivos
                valores = valores[:n_arquivos]
            
            # Preenche os campos de área
            for i, entry in enumerate(self.entradas_area):
                if i < len(valores):
                    entry.delete(0, tk.END)
                    entry.insert(0, f"{valores[i]:.4f}")
            
            popup.destroy()
            self._status(f"{len(valores)} área(s) preenchida(s) automaticamente.")
        
        # Botões
        tk.Button(btn_frame, text="Cancelar", command=popup.destroy,
                  bg="#f1f5f9", fg="#555555", font=("Segoe UI", 9),
                  relief="flat", padx=14, pady=6, cursor="hand2").pack(side="left", padx=4)
        
        tk.Button(btn_frame, text="✅ Preencher Áreas", command=_processar_areas,
                  bg="#0e7490", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=14, pady=6, cursor="hand2").pack(side="left", padx=4)
        
        # Atalho Ctrl+Enter para processar
        text_area.bind("<Control-Return>", lambda e: _processar_areas())
        popup.bind("<Escape>", lambda e: popup.destroy())

    # ── passos ────────────────────────────────────────────────────────────────
    def _step1(self):
        arquivos = filedialog.askopenfilenames(
            title="Selecionar arquivos .dat",
            filetypes=[("Arquivos DAT", "*.dat"), ("Todos", "*.*")]
        )
        if not arquivos:
            return

        self.arquivos_sel = arquivos
        self.app.carregar_arquivos(arquivos)
        self._atualizar_lista()
        self._atualizar_areas()
        self._atualizar_menu_amostras()
        self._set_step("1", SUCCESS, f"{len(arquivos)} arquivo(s) carregado(s)")
        self._set_step("2", TEXT_DIM, "Pendente")
        self._set_step("3", TEXT_DIM, "Pendente")
        self._set_step("4", TEXT_DIM, "Pendente")
        self._status(f"{len(arquivos)} arquivo(s) selecionado(s).")

    def _step2(self):
        pasta = filedialog.askdirectory(title="Selecionar pasta de saída")
        if not pasta:
            return
        self.pasta_saida = pasta
        self._set_step("2", SUCCESS, f"{pasta}")
        self._status(f"Pasta de saída: {pasta}")

    def _step4(self):
        if not self.arquivos_sel:
            self._erro("Selecione os arquivos .dat primeiro (Passo 1).")
            return
        try:
            areas = [float(e.get()) for e in self.entradas_area]
            if any(a <= 0 for a in areas):
                raise ValueError
        except Exception:
            self._erro("Preencha todas as Áreas com valores numéricos positivos.")
            return

        self.app.aplicar_areas(areas)
        self._set_step("3", SUCCESS, "Áreas informadas")
        self._set_step("4", SUCCESS, "Dados processados com sucesso")
        self._status("Dados processados! Você já pode visualizar e exportar.")

        # atualiza painel de informações se houver amostra selecionada
        idx  = self._idx_amostra_ativa()
        if idx is not None:
            self._sincronizar_painel(idx)
        self.grafico.redesenhar()

    # ── exportações ───────────────────────────────────────────────────────────
    def _plotar(self):
        if not self._check_proc():
            return
        ok = self.grafico.redesenhar()
        if not ok:
            self._erro("Processe os dados primeiro (Passo 4).")
        else:
            self._status("Graficos atualizados na interface.")

    def _exportar_png(self):
        if not self._check_proc():
            return
        pasta = self.pasta_saida or filedialog.askdirectory(title="Pasta para salvar PNG")
        if not pasta:
            return
        self.app.exportar_png(pasta)
        self._ok(f"Imagens salvas em:\n{pasta}")

    def _exportar_excel(self):
        if not self._check_proc():
            return
        pasta = self.pasta_saida or filedialog.askdirectory(title="Pasta para salvar Excel")
        if not pasta:
            return
        caminho = self.app.exportar_excel(pasta)
        self._ok(f"Excel gerado:\n{caminho}")

    def _gerar_pdf(self):
        if not self._check_proc():
            return
        pasta = self.pasta_saida or filedialog.askdirectory(title="Pasta para salvar PDF")
        if not pasta:
            return
        caminho = self.app.gerar_pdf(pasta)
        self._ok(f"PDF gerado:\n{caminho}")

    def _toggle_grid(self):
        self.app.mostrar_grid = not self.app.mostrar_grid
        if self.app.mostrar_grid:
            self.btn_grid.config(text="▦  Grid  ON",  bg="#0e7490")
        else:
            self.btn_grid.config(text="▦  Grid  OFF", bg="#374151")
        self.grafico.redesenhar()
        self._status(f"Grid {'ativado' if self.app.mostrar_grid else 'desativado'}.")

    # ── estilo de linha ───────────────────────────────────────────────────────
    def _editar_titulo_pot(self):
        self.grafico.editar_titulo_pot(self.root)

    def _editar_titulo_den(self):
        self.grafico.editar_titulo_den(self.root)

    def _editar_xlabel_pot(self):
        self.grafico.editar_xlabel_pot(self.root)

    def _editar_ylabel_pot(self):
        self.grafico.editar_ylabel_pot(self.root)

    def _editar_xlabel_den(self):
        self.grafico.editar_xlabel_den(self.root)

    def _editar_ylabel_den(self):
        self.grafico.editar_ylabel_den(self.root)

    def _editar_nome_amostra(self):
        idx = self._idx_amostra_ativa()
        if idx is None:
            self._erro("Selecione uma amostra em 'Arquivos selecionados' primeiro.")
            return
        self.grafico.editar_nome_amostra(self.root, idx)
        self._atualizar_menu_amostras()
        self.amostra_var.set(self.app.experimentos[idx].nome)
        self.amostra_idx_sel = idx
        self._atualizar_lista()
        self._sincronizar_painel(idx)

    def _atualizar_menu_amostras(self):
        """Reconstrói o OptionMenu de amostras."""
        menu = self.amostra_menu["menu"]
        menu.delete(0, "end")
        nomes = [e.nome for e in self.app.experimentos]
        if not nomes:
            self.amostra_var.set("—")
            menu.add_command(label="—", command=lambda: self.amostra_var.set("—"))
            self.amostra_idx_sel = None
            return
        for i, nome in enumerate(nomes):
            menu.add_command(label=nome,
                             command=lambda i=i: self._on_amostra_select(i))
        idx = self.amostra_idx_sel if self.amostra_idx_sel is not None else 0
        idx = min(idx, len(nomes) - 1)
        self._selecionar_amostra_idx(idx)

    def _on_amostra_select(self, idx):
        # Recebe o INDICE diretamente (nao o nome). Resolver por nome era
        # fragil: se duas amostras tivessem o mesmo nome (comum quando
        # arquivos .dat tem nomes repetidos/genericos), a selecao sempre
        # "caia" na primeira ocorrencia daquele nome, travando na Amostra 1
        # mesmo clicando em outras amostras no menu.
        if not isinstance(idx, int):
            # placeholder inicial "—" antes de qualquer arquivo carregado
            return
        if 0 <= idx < len(self.app.experimentos):
            self._selecionar_amostra_idx(idx)

    def _idx_amostra_ativa(self):
        # Fonte da verdade e o INDICE selecionado (self.amostra_idx_sel),
        # nao o nome exibido na caixa -- nomes podem se repetir entre
        # amostras diferentes, o que fazia a busca por nome sempre
        # resolver para a primeira ocorrencia.
        if self.amostra_idx_sel is not None and self.amostra_idx_sel < len(self.app.experimentos):
            return self.amostra_idx_sel
        idx = self._idx_amostra(self.amostra_var.get())
        if idx is not None:
            return idx
        return None

    def _selecionar_amostra_idx(self, idx):
        if idx is None or idx >= len(self.app.experimentos):
            return
        self.amostra_idx_sel = idx
        self.amostra_var.set(self.app.experimentos[idx].nome)
        self._sincronizar_painel(idx)
        self._atualizar_lista_selecao()

    def _atualizar_lista_selecao(self):
        for i, row in enumerate(getattr(self, "lista_rows", [])):
            selecionado = i == self.amostra_idx_sel
            bg = ACCENT2 if selecionado else CARD
            fg = "white" if selecionado else TEXT
            row.config(bg=bg)
            for child in row.winfo_children():
                child.config(bg=bg)
                if isinstance(child, tk.Label):
                    child.config(fg=fg if child.cget("text").strip() else ACCENT)

    def _sincronizar_painel(self, idx):
        """Atualiza os controles de estilo e informações com os valores da amostra."""
        e = self.app.experimentos[idx]
        nome_estilo = next((k for k, v in ESTILOS_LINHA.items()
                            if v == e.linestyle), "Sólida")
        self.estilo_var.set(nome_estilo)
        self.lw_var.set(e.linewidth)
        self.lw_label.config(text=f"{e.linewidth:.1f} pt")
        self.cor_preview.config(bg=e.color)

        # informações — só disponíveis após processamento
        if self.app.processado and e.tempo:
            t_total = e.tempo[-1] - e.tempo[0]
            n       = len(e.tempo)
            v       = np.array(e.potencial)
            j       = np.array(e.densidade)
            self.info_var.set(
                f"Nome: {e.nome}\n"
                f"Pontos: {n}   |   Duração: {t_total:.1f} s\n"
                f"Potencial Pico: {np.max(v):.4f} V   |   Final: {v[-1]:.4f} V\n"
                f"Dens. corrente Pico: {np.max(j):.4f} mA/cm²   |   Final: {j[-1]:.4f} mA/cm²"
            )
        else:
            self.info_var.set(f"Nome: {e.nome}\nProcesse os dados (Passo 4) para ver as informações.")

    def _escolher_cor(self):
        from tkinter.colorchooser import askcolor
        idx  = self._idx_amostra_ativa()
        if idx is None:
            return
        nome = self.app.experimentos[idx].nome
        cor_atual = self.app.experimentos[idx].color
        resultado = askcolor(color=cor_atual,
                             title=f"Escolher cor {nome}",
                             parent=self.root)
        if resultado and resultado[1]:
            nova_cor = resultado[1]
            self.app.experimentos[idx].color = nova_cor
            self.cor_preview.config(bg=nova_cor)
            self.grafico.redesenhar()
            self._status(f"Cor de '{nome}' alterada.")

    def _idx_amostra(self, nome):
        for i, e in enumerate(self.app.experimentos):
            if e.nome == nome:
                return i
        return None

    def _on_estilo_change(self, valor):
        idx  = self._idx_amostra_ativa()
        if idx is None:
            return
        nome = self.app.experimentos[idx].nome
        self.app.experimentos[idx].linestyle = ESTILOS_LINHA[valor]
        self.grafico.redesenhar()
        self._status(f"Tipo de linha de '{nome}' alterado para {valor}.")

    def _on_lw_change(self, val):
        idx  = self._idx_amostra_ativa()
        if idx is None:
            return
        lw = float(val)
        self.app.experimentos[idx].linewidth = lw
        self.lw_label.config(text=f"{lw:.1f} pt")
        self.grafico.redesenhar()

    # ── UI helpers ────────────────────────────────────────────────────────────
    def _atualizar_lista(self):
        for w in self.lista_frame.winfo_children():
            w.destroy()
        self.lista_rows = []

        if not self.app.experimentos:
            tk.Label(self.lista_frame, text="Nenhum arquivo carregado.",
                     font=FONT_SMALL, fg=TEXT_DIM, bg=CARD).pack(anchor="w")
            return

        for idx, exp in enumerate(self.app.experimentos):
            row = tk.Frame(self.lista_frame, bg=CARD)
            row.pack(fill="x", pady=2)
            row.config(cursor="hand2")
            self.lista_rows.append(row)

            icon = tk.Label(row, text="  *", font=FONT_BODY, fg=ACCENT, bg=CARD)
            icon.pack(side="left")
            nome = tk.Label(row, text=exp.nome, font=FONT_MONO,
                            fg=TEXT, bg=CARD, anchor="w")
            nome.pack(side="left", fill="x", expand=True, padx=4)

            for widget in (row, icon, nome):
                widget.bind("<Button-1>", lambda e, i=idx: self._selecionar_amostra_idx(i))
                widget.bind("<Double-Button-1>", lambda e, i=idx: self._editar_nome_amostra_idx(i))

        self._atualizar_lista_selecao()

    def _editar_nome_amostra_idx(self, idx):
        if idx is None or idx >= len(self.app.experimentos):
            return
        self.amostra_idx_sel = idx
        self.grafico.editar_nome_amostra(self.root, idx)
        self._atualizar_menu_amostras()
        self.amostra_var.set(self.app.experimentos[idx].nome)
        self._atualizar_lista()
        self._sincronizar_painel(idx)

    def _atualizar_areas(self):
        for w in self.areas_frame.winfo_children():
            w.destroy()
        self.entradas_area = []

        if not self.arquivos_sel:
            lbl = tk.Label(self.areas_frame, text="Nenhum arquivo carregado.",
                           font=FONT_SMALL, fg=TEXT_DIM, bg=SURFACE)
            lbl.pack(padx=14, pady=10)
            self._vincular_scroll_areas_widget(lbl)
            return

        header = tk.Frame(self.areas_frame, bg=SURFACE)
        header.pack(fill="x", padx=8, pady=(8, 4))
        self._vincular_scroll_areas_widget(header)
        lbl_arq = tk.Label(header, text="Arquivo", font=("Segoe UI", 9, "bold"),
                           fg=TEXT_DIM, bg=SURFACE, width=18, anchor="w")
        lbl_arq.grid(row=0, column=0, padx=4)
        lbl_area = tk.Label(header, text="Área (cm²)", font=("Segoe UI", 9, "bold"),
                            fg=TEXT_DIM, bg=SURFACE)
        lbl_area.grid(row=0, column=1, padx=8)
        self._vincular_scroll_areas_widget(lbl_arq)
        self._vincular_scroll_areas_widget(lbl_area)

        separator(self.areas_frame).pack(fill="x", padx=8, pady=2)

        for i, arq in enumerate(self.arquivos_sel):
            row = tk.Frame(self.areas_frame, bg=SURFACE if i % 2 == 0 else CARD)
            row.pack(fill="x", padx=8, pady=1)
            self._vincular_scroll_areas_widget(row)

            nome = os.path.basename(arq).replace(".dat", "")
            nome_lbl = tk.Label(row, text=nome, font=FONT_MONO, fg=TEXT,
                                bg=row["bg"], width=20, anchor="w")
            nome_lbl.grid(row=0, column=0, padx=6, pady=4)
            self._vincular_scroll_areas_widget(nome_lbl)

            entry = tk.Entry(row, font=FONT_BODY, bg=BG, fg=TEXT,
                             insertbackground=TEXT, relief="flat",
                             highlightthickness=1, highlightcolor=ACCENT,
                             highlightbackground=BORDER, width=8)
            entry.grid(row=0, column=1, padx=8, pady=4)
            self._vincular_scroll_areas_widget(entry)
            self.entradas_area.append(entry)

        self.root.after_idle(
            lambda: self.scroll_canvas.configure(scrollregion=self.scroll_canvas.bbox("all"))
        )

    def _set_step(self, num, cor, texto):
        self.step_labels[num].config(text=texto, fg=cor)

    def _status(self, msg):
        self.status_var.set(msg)

    def _check_proc(self):
        if not self.app.processado:
            self._erro("Processe os dados primeiro (Passo 4).")
            return False
        return True

    def _erro(self, msg):
        messagebox.showerror("H-AnodPlot", msg)

    def _ok(self, msg):
        messagebox.showinfo("H-AnodPlot", msg)


# =============================================================================
# ENTRY POINT
# =============================================================================
if __name__ == "__main__":
    App()